import time
import random
import os
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage, ImageOps
import datetime

# Extract unique library IDs and ad blocks
def extract_library_ids(driver):
    library_data = []
    elements = driver.find_elements(By.XPATH, "//span[contains(text(),'Library ID')]")
    for element in elements:
        try:
            text = element.text
            if text.startswith("Library ID:"):
                library_id = text.split(":")[1].strip()
                parent_div = element.find_element(By.XPATH, "ancestor::div[7]")
                library_data.append((library_id, parent_div))
        except Exception as e:
            print(f"Error extracting Library ID: {e}")
    return library_data

# Generate Facebook ad links
def generate_library_urls(library_ids):
    return [f"https://www.facebook.com/ads/library/?id={library_id}" for library_id in library_ids]

# Scroll to bottom with logic to wait for new content
def scroll_until_end(driver, max_scrolls=1000):
    seen_ids = set()
    scroll_count = 0

    while scroll_count < max_scrolls:
        library_data = extract_library_ids(driver)
        new_ids = {lib_id for lib_id, _ in library_data} - seen_ids

        if not new_ids:
            print("No new ads found. Stopping scroll.")
            break

        seen_ids.update(new_ids)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(random.uniform(2.5, 4.5))
        scroll_count += 1

    print(f"Finished scrolling after {scroll_count} scrolls.")

# Capture ad screenshots from freshly extracted library data
def capture_all_ads_from_fresh(driver, folder_path, library_data):
    seen_ids = set()

    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    driver.execute_script("document.body.style.zoom='67%'")

    captured_ids = []
    for lib_id, ad_element in library_data:
        if lib_id in seen_ids:
            continue
        try:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", ad_element)
            WebDriverWait(driver, 10).until(
                EC.visibility_of(ad_element)
            )
            time.sleep(1.5)

            screenshot_path = os.path.join(folder_path, f"{lib_id}_raw.png")
            ad_element.screenshot(screenshot_path)

            with PILImage.open(screenshot_path) as im:
                width, height = im.size
                crop_right = width * 0.75
                crop_bottom = height * 0.70
                cropped_image = im.crop((0, 0, int(crop_right), int(crop_bottom)))
                bordered = ImageOps.expand(cropped_image, border=3, fill='red')
                final_path = os.path.join(folder_path, f"{lib_id}.png")
                bordered.save(final_path)

            os.remove(screenshot_path)
            captured_ids.append(lib_id)
            seen_ids.add(lib_id)
            print(f"Saved: {lib_id}")
        except Exception as e:
            print(f"Error on {lib_id}: {e}")

    return captured_ids

# Insert images into Excel
def insert_images_in_excel(excel_path, folder_path, library_ids):
    wb = load_workbook(excel_path)
    ws = wb.active
    ws.cell(row=4, column=4).value = "Screenshot"  # Header in row 4

    for i, lib_id in enumerate(library_ids, start=5):  # Start inserting from row 5
        img_path = os.path.join(folder_path, f"{lib_id}.png")
        if os.path.exists(img_path):
            try:
                excel_img = ExcelImage(img_path)
                ws.add_image(excel_img, f"D{i}")
            except Exception as e:
                print(f"Error inserting image for {lib_id}: {e}")

    wb.save(excel_path)
    print(f"Images embedded in Excel: {excel_path}")

# Main
def main():
    facebook_ads_link = input("Enter the Facebook Ads link: ").strip()
    folder_path = input("Enter the folder path where you want to save the output: ").strip()
    keyword = input("Enter the keyword (e.g., Aviator): ").strip()

    edge_options = Options()
    edge_options.add_argument("--headless")
    edge_options.add_argument("--disable-gpu")
    edge_options.add_argument("--window-size=1600,6000")
    edge_options.add_argument("--disable-blink-features=AutomationControlled")

    service = Service("/home/vassu/.wdm/drivers/edgedriver/linux64/135.0.3179.85/msedgedriver")
    driver = webdriver.Edge(service=service, options=edge_options)

    try:
        driver.get(facebook_ads_link)
        time.sleep(random.uniform(8, 12))

        print("Scrolling page to load all ads...")
        scroll_until_end(driver)

        print("Extracting fresh ads after scrolling...")
        library_data_after_scroll = extract_library_ids(driver)

        print("Capturing all ad screenshots...")
        captured_ids = capture_all_ads_from_fresh(driver, folder_path, library_data_after_scroll)
        library_urls = generate_library_urls(captured_ids)

        df = pd.DataFrame({
            "S.No.": range(1, len(captured_ids) + 1),
            "Library ID": captured_ids,
            "Library ID URL": library_urls
        })

        excel_path = os.path.join(folder_path, "facebook_ads_library_ids.xlsx")
        df.to_excel(excel_path, index=False, startrow=3)  # Data starts at row 4

        # Add keyword, URL, and script running date at the top
        wb = load_workbook(excel_path)
        ws = wb.active
        ws.cell(row=1, column=1).value = f"Keyword: {keyword}"
        ws.cell(row=2, column=1).value = f"URL: {facebook_ads_link}"
        ws.cell(row=3, column=1).value = f"Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        wb.save(excel_path)

        insert_images_in_excel(excel_path, folder_path, captured_ids)

    finally:
        driver.quit()
        print("Browser closed.")

if __name__ == "__main__":
    main()

